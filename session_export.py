"""
session_export.py
-----------------
Phase 1 of the session EXPORTS layer: write a workbench session's RESULTS into its
`{name}/exports/` folder as durable, portable artifacts — REUSING the data already in
`variant_model` ResultSlots / ChainDesign (NO recompute).

ONE row-building pass (`build_tables`) walks DesignSession → ChainDesign → Variant and
reshapes the existing dicts into long-format tables; `export_session` then feeds the SAME
tables to BOTH writers:
  exports/results.xlsx        — one SHEET per result type (Summary first)
  exports/csv/{type}.csv      — one CSV per result type, same columns as its sheet

FAIL-LOUD: a result type with no data is SKIPPED entirely (no empty/header-only sheet or
CSV); if nothing has data, no files are written and the caller reports "nothing to export".
The SUMMARY tab joins across result types that may be only PARTIALLY populated — a variant
missing a given result yields a clean BLANK cell (None), never 0 and never a crash.

Pure / GUI-agnostic / testable: takes `design_sessions` (the dict blobs SessionState persists)
+ an output dir; returns a {written, skipped, files} report.
"""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Canonical result types → (sheet title, csv slug). Order = sheet order in the workbook.
_TYPES: List[Tuple[str, str, str]] = [
    ("summary",               "Summary",               "summary"),
    ("sequences",             "Sequences",             "sequences"),
    ("substitutions",         "Substitutions",         "substitutions"),
    ("fold_plddt",            "Fold pLDDT",             "fold_plddt"),
    ("deviation",             "Deviation",              "deviation"),
    ("stability_ddg",         "Stability ddG",          "stability_ddg"),
    ("solubility",            "Solubility",             "solubility"),
    ("template_assist",       "Template assist",        "template_assist"),
    ("template_assist_dflex", "Template assist dflex",  "template_assist_dflex"),
    ("structural_align",      "Structural align",       "structural_align"),
]
_TITLE = {k: t for k, t, _s in _TYPES}
_SLUG = {k: s for k, _t, s in _TYPES}
# the data types that count as "exportable content" (Summary is a derived roll-up, never counts);
# sequences/substitutions are DESIGN content (phase 1.5) — a bare construct with no results still
# exports its sequence, so they count toward "is there anything to export".
_DATA_KEYS = [k for k, _t, _s in _TYPES if k != "summary"]

_COLUMNS: Dict[str, List[str]] = {
    "summary": ["model", "design_chain", "row", "n_mutations", "mean_plddt", "sum_ddg",
                "solubility_delta", "max_dRMSD", "adoption", "tm_align"],
    "sequences": ["model", "design_chain", "row", "source", "length", "sequence",
                  "mpnn_score", "mpnn_recovery"],
    "substitutions": ["model", "design_chain", "variant", "kind", "resnum", "from_aa", "to_aa",
                      "residues", "source", "score", "recovery", "recommendation"],
    "fold_plddt": ["model", "design_chain", "row", "engine", "target", "resnum", "plddt"],
    "deviation": ["model", "design_chain", "variant", "chain", "resnum",
                  "dRMSD", "dRMSD_floor", "lDDT", "lDDT_floor"],
    "stability_ddg": ["model", "design_chain", "variant", "resnum", "from_aa", "to_aa",
                      "ddg", "ddg_source", "combined_score", "recommendation", "sum_ddg", "tier"],
    "solubility": ["model", "design_chain", "variant", "camsol_variant", "camsol_wt", "delta"],
    "template_assist": ["model", "design_chain", "template_label", "unguided_mean_plddt",
                        "guided_mean_plddt", "d_plddt", "n_stabilized", "n_residues",
                        "mean_d_flex", "max_adoption", "tm_adopt", "force", "threshold"],
    "template_assist_dflex": ["model", "design_chain", "resnum", "d_flex"],
    "structural_align": ["model", "design_chain", "reference", "ref_label", "tm_ref",
                         "tm_query", "rmsd", "n_aligned", "norm"],
}


def _as_int(k: Any) -> Optional[int]:
    try:
        return int(k)
    except (TypeError, ValueError):
        return None


def _sorted_resnum_items(m: Dict[Any, Any]):
    """(resnum-key, value) pairs ordered by numeric resnum where possible (keys may be int OR
    str after a JSON round-trip), stable for non-numeric keys."""
    return sorted((m or {}).items(), key=lambda kv: (_as_int(kv[0]) is None, _as_int(kv[0]) or 0, str(kv[0])))


def _split_dev_key(k: Any) -> Tuple[Optional[str], Any]:
    """A deviation key is `resnum` (monomer) or `"chain:resno"` (multichain)."""
    s = str(k)
    if ":" in s:
        c, r = s.split(":", 1)
        return c, r
    return None, s


def _seq_from_cells(cells) -> str:
    """The ungapped sequence from a cells list ([{col,resnum,aa}], aa None = gap)."""
    return "".join(c.get("aa") for c in (cells or []) if c.get("aa"))


def _mpnn_score(prov: Dict[str, Any], model: str, var_seq: str,
                ppr: Optional[Dict[str, Any]]) -> Tuple[Any, Any]:
    """(score, recovery) for an MPNN-imported variant — joined from `proteinmpnn_results[model]
    ["data"]["sequences"][design_k]` (the design model_id IS the design_sessions key; design_k is
    the enumerate index `import_mpnn_designs` recorded) and VERIFIED by SEQUENCE match, so an
    overwritten/later run (only the latest is stored per model) never attaches a WRONG score —
    it blanks instead. (None, None) when no proteinmpnn_results given / no clean match."""
    if not ppr:
        return None, None
    dk = prov.get("design_k")
    if not isinstance(dk, int) or dk < 0:
        return None, None
    seqs = ((ppr.get(model) or {}).get("data") or {}).get("sequences") or []
    if dk >= len(seqs):
        return None, None
    s = seqs[dk] or {}
    if s.get("sequence") and var_seq and s["sequence"] != var_seq:
        return None, None                              # stored result is a different run → blank
    return s.get("score"), s.get("recovery")


def _plddt_rows(model: str, chain: str, row_label: str, fold: Dict[str, Any]) -> List[Dict[str, Any]]:
    plddt = (fold or {}).get("plddt") or {}
    if not plddt:
        return []
    eng, tgt = fold.get("engine"), fold.get("target")
    return [{"model": model, "design_chain": chain, "row": row_label, "engine": eng,
             "target": tgt, "resnum": _as_int(rn) if _as_int(rn) is not None else rn,
             "plddt": val}
            for rn, val in _sorted_resnum_items(plddt)]


def build_tables(design_sessions: Dict[str, Any],
                 proteinmpnn_results: Optional[Dict[str, Any]] = None) -> Dict[str, List[Dict[str, Any]]]:
    """ONE pass over the design tree → {type_key: [row-dict, …]}. Only the rows; columns come
    from `_COLUMNS`. Reshapes existing data only (no recompute). The Summary roll-up joins
    per-variant across result types, leaving a BLANK (None) cell where a result is absent.
    *proteinmpnn_results* (optional) lets MPNN substitution/sequence rows carry the design's
    score + recovery (verified by sequence match; see `_mpnn_score`)."""
    rows: Dict[str, List[Dict[str, Any]]] = {k: [] for k, _t, _s in _TYPES}

    for model, ds in (design_sessions or {}).items():
        model = str(model)
        for cd in ((ds or {}).get("chains") or {}).values():
            chain = cd.get("rep_chain") or "?"
            tf = cd.get("template_fold") or {}
            gf = cd.get("guided_fold") or {}
            ta = cd.get("template_assist") or {}
            sa = cd.get("structural_align") or {}

            # ── design CONTENT: the template (T) sequence (phase 1.5) ──
            tseq = _seq_from_cells(cd.get("template_cells"))
            if tseq:
                rows["sequences"].append({"model": model, "design_chain": chain, "row": "T",
                                          "source": "template", "length": len(tseq), "sequence": tseq,
                                          "mpnn_score": None, "mpnn_recovery": None})

            # ── construct-level (per ChainDesign) ──
            rows["fold_plddt"] += _plddt_rows(model, chain, "T", tf)
            rows["fold_plddt"] += _plddt_rows(model, chain, "guided", gf)
            if ta:
                rows["template_assist"].append({
                    "model": model, "design_chain": chain, "template_label": ta.get("template_label"),
                    "unguided_mean_plddt": ta.get("unguided_mean_plddt"),
                    "guided_mean_plddt": ta.get("guided_mean_plddt"), "d_plddt": ta.get("d_plddt"),
                    "n_stabilized": ta.get("n_stabilized"), "n_residues": ta.get("n_residues"),
                    "mean_d_flex": ta.get("mean_d_flex"), "max_adoption": ta.get("max_adoption"),
                    "tm_adopt": ta.get("tm_adopt"), "force": ta.get("force"),
                    "threshold": ta.get("threshold")})
                for rn, val in _sorted_resnum_items(ta.get("d_flex") or {}):
                    rows["template_assist_dflex"].append({
                        "model": model, "design_chain": chain,
                        "resnum": _as_int(rn) if _as_int(rn) is not None else rn, "d_flex": val})
            if sa:
                rows["structural_align"].append({
                    "model": model, "design_chain": chain, "reference": sa.get("reference"),
                    "ref_label": sa.get("ref_label"), "tm_ref": sa.get("tm_ref"),
                    "tm_query": sa.get("tm_query"), "rmsd": sa.get("rmsd"),
                    "n_aligned": sa.get("n_aligned"), "norm": sa.get("norm")})

            # construct-level Summary roll-up values (shared by every variant of this cd)
            adoption = ta.get("max_adoption") if ta else None
            if adoption is None and ta:
                adoption = ta.get("tm_adopt")
            tm_align = sa.get("tm_ref") if sa else None
            # a "T" summary row so a construct with assist/align but no variants still appears
            if tf.get("plddt") or ta or sa:
                rows["summary"].append({
                    "model": model, "design_chain": chain, "row": "T", "n_mutations": 0,
                    "mean_plddt": tf.get("mean_plddt"), "sum_ddg": None, "solubility_delta": None,
                    "max_dRMSD": None, "adoption": adoption, "tm_align": tm_align})

            # ── per-variant ──
            for v in cd.get("variants") or []:
                vid = v.get("id")
                vsource = v.get("source")
                prov = v.get("provenance") or {}

                # ── design CONTENT (phase 1.5): variant sequence + per-change substitution rows ──
                vseq = _seq_from_cells(v.get("cells"))
                mscore, mrec = (_mpnn_score(prov, model, vseq, proteinmpnn_results)
                                if vsource == "proteinmpnn" else (None, None))
                if vseq:
                    rows["sequences"].append({"model": model, "design_chain": chain, "row": vid,
                                              "source": vsource, "length": len(vseq), "sequence": vseq,
                                              "mpnn_score": mscore, "mpnn_recovery": mrec})
                accepted = {a.get("resnum"): a for a in (prov.get("accepted") or [])}
                for m in v.get("mutations") or []:
                    rn = m.get("resnum")
                    if vsource == "proteinmpnn":
                        score, rec, recd = mscore, mrec, None
                    else:                                  # accepted-suggestion subs carry a per-resnum score
                        acc = accepted.get(rn) or {}
                        score, rec, recd = acc.get("combined_score"), None, acc.get("recommendation")
                    rows["substitutions"].append({
                        "model": model, "design_chain": chain, "variant": vid, "kind": "substitution",
                        "resnum": rn, "from_aa": m.get("from_aa"), "to_aa": m.get("to_aa"),
                        "residues": None, "source": m.get("source"), "score": score,
                        "recovery": rec, "recommendation": recd})
                for e in v.get("indels") or []:
                    rows["substitutions"].append({
                        "model": model, "design_chain": chain, "variant": vid,
                        "kind": e.get("kind"), "resnum": e.get("resnum"), "from_aa": e.get("from_aa"),
                        "to_aa": None, "residues": e.get("residues"), "source": vsource,
                        "score": mscore if vsource == "proteinmpnn" else None,
                        "recovery": mrec if vsource == "proteinmpnn" else None, "recommendation": None})

                res = v.get("results") or {}
                vfold = res.get("fold") or {}
                stab = res.get("stability") or {}
                sol = res.get("solubility") or {}
                dev = vfold.get("deviation") or {}

                rows["fold_plddt"] += _plddt_rows(model, chain, vid, vfold)

                max_drmsd = None
                if dev:
                    ddm = dev.get("ddm") or {}
                    lddt = dev.get("lddt") or {}
                    fddm = dev.get("floor_ddm") or {}
                    flddt = dev.get("floor_lddt") or {}
                    dvals = [x for x in ddm.values() if isinstance(x, (int, float))]
                    max_drmsd = max(dvals) if dvals else None
                    seen = set()
                    for k in list(ddm.keys()) + [k for k in lddt.keys() if k not in ddm]:
                        if k in seen:
                            continue
                        seen.add(k)
                        ch, rn = _split_dev_key(k)
                        rows["deviation"].append({
                            "model": model, "design_chain": chain, "variant": vid,
                            "chain": ch or chain, "resnum": _as_int(rn) if _as_int(rn) is not None else rn,
                            "dRMSD": ddm.get(k), "dRMSD_floor": fddm.get(k),
                            "lDDT": lddt.get(k), "lDDT_floor": flddt.get(k)})

                if stab.get("rows"):
                    for r in stab["rows"]:
                        rows["stability_ddg"].append({
                            "model": model, "design_chain": chain, "variant": vid,
                            "resnum": r.get("resnum"), "from_aa": r.get("from_aa"),
                            "to_aa": r.get("to_aa"), "ddg": r.get("ddg"),
                            "ddg_source": r.get("ddg_source"),
                            "combined_score": r.get("combined_score"),
                            "recommendation": r.get("recommendation"),
                            "sum_ddg": stab.get("sum_ddg"), "tier": stab.get("tier")})

                if sol:
                    rows["solubility"].append({
                        "model": model, "design_chain": chain, "variant": vid,
                        "camsol_variant": sol.get("variant"), "camsol_wt": sol.get("wt"),
                        "delta": sol.get("delta")})

                # per-variant Summary roll-up — blank (None) where a result is absent
                if vfold or stab or sol or dev:
                    rows["summary"].append({
                        "model": model, "design_chain": chain, "row": vid,
                        "n_mutations": len(v.get("mutations") or []),
                        "mean_plddt": vfold.get("mean_plddt"),
                        "sum_ddg": stab.get("sum_ddg") if stab else None,
                        "solubility_delta": sol.get("delta") if sol else None,
                        "max_dRMSD": max_drmsd, "adoption": adoption, "tm_align": tm_align})

    return rows


def export_session(design_sessions: Dict[str, Any], exports_dir,
                   proteinmpnn_results: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """Write `results.xlsx` (one sheet per non-empty type) + `csv/{type}.csv` (+ `sequences.fasta`)
    into *exports_dir*. FAIL-LOUD: a type with no rows is SKIPPED (no empty file); if NOTHING has
    data (no design content nor results), nothing is written (`any=False`). Returns
    {any, written, skipped, files}."""
    exports_dir = Path(exports_dir)
    tables = build_tables(design_sessions, proteinmpnn_results)
    present = [k for k in _DATA_KEYS if tables.get(k)]             # data types (incl. sequences) WITH rows
    skipped = [_TITLE[k] for k in _DATA_KEYS if not tables.get(k)]
    if not present:
        return {"any": False, "written": [], "skipped": [_TITLE[k] for k in _DATA_KEYS], "files": []}

    emit = [k for k, _t, _s in _TYPES if tables.get(k)]            # canonical order; Summary iff it has rows
    exports_dir.mkdir(parents=True, exist_ok=True)
    files: List[str] = []

    # ── workbook (one sheet per emitted type, in canonical order) ──
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    for key, title, _slug in _TYPES:
        if key not in emit:
            continue
        cols = _COLUMNS[key]
        ws = wb.create_sheet(title=title[:31])                    # Excel 31-char sheet-name cap
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            ws.cell(row=1, column=c).font = bold
        ws.freeze_panes = "A2"
        for r in tables[key]:
            ws.append([r.get(c) for c in cols])
    xlsx_path = exports_dir / "results.xlsx"
    wb.save(str(xlsx_path))
    files.append(str(xlsx_path))

    # ── per-type CSVs (same columns) ──
    csv_dir = exports_dir / "csv"
    csv_dir.mkdir(parents=True, exist_ok=True)
    for key in emit:
        cols = _COLUMNS[key]
        p = csv_dir / f"{_SLUG[key]}.csv"
        with open(p, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(cols)
            for r in tables[key]:
                w.writerow([r.get(c) for c in cols])
        files.append(str(p))

    # ── sequences.fasta (the natural interchange) — T + every variant ──
    if tables.get("sequences"):
        fasta = exports_dir / "sequences.fasta"
        with open(fasta, "w", encoding="utf-8") as fh:
            for r in tables["sequences"]:
                hdr = f">{r['model']}_{r['design_chain']}_{r['row']} source={r.get('source')}"
                if r.get("mpnn_score") is not None:
                    hdr += f" mpnn_score={r['mpnn_score']} recovery={r.get('mpnn_recovery')}"
                fh.write(hdr + "\n" + (r.get("sequence") or "") + "\n")
        files.append(str(fasta))

    # ── phase 2: relevance-gated PROFILE figures (a bonus layer — a plotting failure never breaks
    # the data export that already succeeded). Only per-residue profile types render a plot. ──
    figures = {"written": [], "skipped": [], "error": None}
    try:
        import session_figures
        figures = session_figures.render_profile_figures(tables, exports_dir / "figures")
        files.extend(str(exports_dir / "figures" / n) for n in figures["written"])
    except Exception as exc:
        figures["error"] = f"{type(exc).__name__}: {exc}"

    return {"any": True, "written": [_TITLE[k] for k in emit], "skipped": skipped,
            "files": files, "figures": figures}
