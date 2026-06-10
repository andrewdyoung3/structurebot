"""
scripts/stability_export.py — dual-output (Raw + Calibrated) candidate exporter.

Reads the lossless data-gen JSONL (scripts/stability_datagen.py output) and emits a
per-candidate spreadsheet with, for every ddG voter, BOTH:
  • a RAW / UNMODIFIED ddG column — the lossless source of truth, NEVER altered; and
  • a CALIBRATED ddG column — a DERIVED view.

CALIBRATION IS AN IDENTITY PASSTHROUGH FOR NOW.  No calibration fit exists yet (the
§9 calibration benchmark is unexecuted), so calibrated == raw and the calibration
metadata columns are empty placeholders — structured so a real per-voter fit can
later fill offset / slope / uncertainty / provenance / version / OOD-caveat WITHOUT
touching the raw columns.

Property axes (CamSol solubility, ESM fitness) are NOT ddG and live in their own
clearly-separated column group — never mixed into the stability ddG columns.

Excel via openpyxl when available; otherwise a lossless CSV with identical columns.
(The xlsx skill at /mnt/skills/public/xlsx/SKILL.md was not present in this
environment — openpyxl is used directly; CSV is the dependency-free fallback.)

Usage:
  python scripts/stability_export.py --in cache/stability_datagen/rows.jsonl \
                                     --out cache/stability_datagen/candidates.xlsx
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Windows UTF-8 stdout convention (§5): headers/log lines use →/— and §.
if sys.platform == "win32" and hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

sys.path.insert(0, str(Path(__file__).resolve().parent))
from rosetta_clash import rosetta_confidence  # noqa: E402

# Calibration is identity for now; this tag versions the (non-)transform so a later
# recalibration never silently reinterprets an old export.
CAL_VERSION = "identity_passthrough_v0"
CAL_STATUS = "UNCALIBRATED — raw passthrough; no calibration fit exists yet (§9 pending)"

DDG_VOTERS = ["rosetta", "rasp", "thermompnn", "dynamut2"]   # stability ddG axes


def calibrate(voter: str, raw: Optional[float]) -> Tuple[Optional[float], Dict[str, Any]]:
    """Placeholder calibration: IDENTITY.  Returns (calibrated_value, meta).

    LOSSLESS — never mutates `raw`.  When a real fit lands, this maps raw→calibrated
    per voter (and possibly per protein) and fills the meta fields; until then it is a
    pass-through so the Calibrated column equals the Raw column and is clearly tagged
    uncalibrated.  Property axes never pass through here (they are not ddG)."""
    meta = {
        "version": CAL_VERSION,
        "offset": None,        # additive correction (none yet)
        "slope": None,         # multiplicative correction (none yet)
        "uncertainty": None,   # from cross-protein spread (none yet)
        "provenance": "uncalibrated",
        "ood_caveat": None,    # out-of-distribution warning (none yet)
    }
    return raw, meta


# ── column model ─────────────────────────────────────────────────────────────────
# (group, column) pairs — group drives the merged top-header band so the Raw /
# Calibrated / Property / Provenance separation is visually unambiguous.

_IDENTITY = ["key", "set", "role", "pdbid", "chain", "resnum", "wt", "mut", "variant"]
_LABELS = ["exp_ddg", "rosetta_ref_ddg"]
_CAL_META = ["cal_version", "cal_status", "cal_offset", "cal_slope",
             "cal_uncertainty", "cal_ood_caveat"]
_PROPERTY = ["property_camsol_solubility", "property_esm_fitness_tolerance"]
_PROVENANCE = ["prov_rosetta", "prov_rasp", "prov_thermompnn", "prov_dynamut2"]
_QC = ["rosetta_confidence"]   # clash-artifact flag (raw value untouched)
_FLAGS = ["rosetta_in_subset", "dynamut2_in_subset"]


def _columns() -> List[Tuple[str, str]]:
    cols: List[Tuple[str, str]] = []
    cols += [("identity", c) for c in _IDENTITY]
    cols += [("experiment", c) for c in _LABELS]
    cols += [("ddG RAW (lossless — never altered)", f"{v}_ddg_raw") for v in DDG_VOTERS]
    cols += [("ddG CALIBRATED (derived; identity now)", f"{v}_ddg_cal") for v in DDG_VOTERS]
    cols += [("calibration metadata", c) for c in _CAL_META]
    cols += [("property axes (NOT ddG)", c) for c in _PROPERTY]
    cols += [("provenance (per-voter training-disjointness)", c) for c in _PROVENANCE]
    cols += [("QC flags (raw never altered)", c) for c in _QC]
    cols += [("subset flags", c) for c in _FLAGS]
    return cols


def _row_values(rec: Dict[str, Any]) -> Dict[str, Any]:
    """Project one lossless JSONL record onto the export columns.  Raw ddG columns
    are copied verbatim (lossless); calibrated columns come from calibrate()."""
    out: Dict[str, Any] = {}
    for c in _IDENTITY:
        out[c] = rec.get(c)
    for c in _LABELS:
        out[c] = rec.get(c)
    # RAW (verbatim) + CALIBRATED (derived) per voter
    cal_meta_seen: Dict[str, Any] = {}
    for v in DDG_VOTERS:
        raw = rec.get(f"{v}_ddg")
        out[f"{v}_ddg_raw"] = raw
        cal, meta = calibrate(v, raw)
        out[f"{v}_ddg_cal"] = cal
        cal_meta_seen = meta  # identical across voters under identity; carried once
    out["cal_version"] = cal_meta_seen.get("version")
    out["cal_status"] = CAL_STATUS
    out["cal_offset"] = cal_meta_seen.get("offset")
    out["cal_slope"] = cal_meta_seen.get("slope")
    out["cal_uncertainty"] = cal_meta_seen.get("uncertainty")
    out["cal_ood_caveat"] = cal_meta_seen.get("ood_caveat")
    # property axes — separated, labelled, never folded into ddG
    out["property_camsol_solubility"] = rec.get("camsol_score")
    out["property_esm_fitness_tolerance"] = rec.get("esm_tolerance")
    for c in _PROVENANCE:
        out[c] = rec.get(c)
    # QC: Rosetta clash-artifact flag — DERIVED from the raw value, which is left
    # untouched in rosetta_ddg_raw above (flag, never delete/alter).
    out["rosetta_confidence"] = rosetta_confidence(rec.get("rosetta_ddg"))
    for c in _FLAGS:
        out[c] = rec.get(c)
    return out


def load_records(jsonl_path: str) -> List[Dict[str, Any]]:
    p = Path(jsonl_path)
    if not p.is_file():
        return []
    out = []
    for ln in p.read_text(encoding="utf-8").splitlines():
        ln = ln.strip()
        if ln:
            out.append(json.loads(ln))
    return out


def export(jsonl_path: str, out_path: str, log=print) -> int:
    records = load_records(jsonl_path)
    cols = _columns()
    rows = [_row_values(r) for r in records]
    suffix = Path(out_path).suffix.lower()
    if suffix == ".xlsx":
        _write_xlsx(cols, rows, out_path)
    else:
        _write_csv(cols, rows, out_path)
    log(f"exported {len(rows)} candidates × {len(cols)} columns → {out_path}")
    return len(rows)


def _write_csv(cols: List[Tuple[str, str]], rows: List[Dict[str, Any]], out_path: str) -> None:
    names = [c for _g, c in cols]
    with open(out_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow([g for g, _c in cols])   # group band
        w.writerow(names)                    # column names
        for r in rows:
            w.writerow([r.get(c) for c in names])


def _write_xlsx(cols: List[Tuple[str, str]], rows: List[Dict[str, Any]], out_path: str) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "candidates"
    names = [c for _g, c in cols]

    # row 1 = merged group band; row 2 = column names
    bold = Font(bold=True)
    grp_font = Font(bold=True, color="FFFFFF")
    fills = {
        "ddG RAW (lossless — never altered)": "1F4E78",
        "ddG CALIBRATED (derived; identity now)": "2E75B6",
        "calibration metadata": "9DC3E6",
        "property axes (NOT ddG)": "548235",
        "provenance (per-voter training-disjointness)": "7F6000",
    }
    # build contiguous group spans
    c0 = 0
    while c0 < len(cols):
        g = cols[c0][0]
        c1 = c0
        while c1 + 1 < len(cols) and cols[c1 + 1][0] == g:
            c1 += 1
        ws.cell(row=1, column=c0 + 1, value=g)
        if c1 > c0:
            ws.merge_cells(start_row=1, start_column=c0 + 1, end_row=1, end_column=c1 + 1)
        cell = ws.cell(row=1, column=c0 + 1)
        cell.alignment = Alignment(horizontal="center")
        if g in fills:
            cell.font = grp_font
            for cc in range(c0 + 1, c1 + 2):
                ws.cell(row=1, column=cc).fill = PatternFill(
                    "solid", fgColor=fills[g])
        else:
            cell.font = bold
        c0 = c1 + 1

    for j, nm in enumerate(names, 1):
        cell = ws.cell(row=2, column=j, value=nm)
        cell.font = bold
    for i, r in enumerate(rows, 3):
        for j, nm in enumerate(names, 1):
            ws.cell(row=i, column=j, value=r.get(nm))
    ws.freeze_panes = "A3"
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp",
                    default="cache/stability_datagen/rows.jsonl")
    ap.add_argument("--out",
                    default="cache/stability_datagen/candidates.xlsx")
    a = ap.parse_args()
    export(a.inp, a.out)


if __name__ == "__main__":
    main()
