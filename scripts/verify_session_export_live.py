"""
VERIFY (no GPU, no ChimeraX) — the session EXPORTS layer + Save As end-to-end on a MIXED session
built through the REAL model path (build_design_session_from_sequence + the real fold_summary /
stability_summary reducers + DesignSession.to_dict), so it catches any drift between what the panel
actually PERSISTS and what session_export reads. Exercises the partial-data Summary + skip-empty +
the self-contained fork — not an all-populated happy case.

Mixed design: a construct with template_assist + structural_align; V1 folded + deviation + stability;
V2 solubility-only; V3 no results at all.

Run:  venv/Scripts/python.exe scripts/verify_session_export_live.py
"""
import sys, tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
try: sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception: pass

import config
from session_state import SessionState
from variant_model import build_design_session_from_sequence, fold_summary, stability_summary
import session_io, session_export

PASS, FAIL = [], []
def check(name, ok): (PASS if ok else FAIL).append(name); print(("  OK  " if ok else " FAIL ") + name)

SEQ = "ACDEFGHIKL"           # 10-mer
author = list(range(1, len(SEQ) + 1))

# ── build a MIXED design via the real builders/reducers ──
design = build_design_session_from_sequence("mix", [(SEQ, 1)])
cd = next(iter(design.chains.values()))

# a real fold CIF so the durable folds/ copy path is exercised end-to-end
_cif = Path(tempfile.gettempdir()) / "boltz_pred_exportverify.cif"
_cif.write_text("data_x\n# fake fold\n")

# construct-level: real fold_summary for T + template_assist + structural_align
cd.template_fold = fold_summary({"engine": "boltz", "target": "monomer", "model_id": "3",
                                 "mean_plddt": 90.0, "plddt": {i: 90.0 - i for i in range(1, 11)},
                                 "cif_path": str(_cif)},
                                author)
cd.template_assist = {"template_label": "8UB2", "unguided_mean_plddt": 92.0, "guided_mean_plddt": 95.0,
                      "d_plddt": 3.0, "n_stabilized": 4, "n_residues": 10, "mean_d_flex": 0.12,
                      "max_adoption": 0.93, "tm_adopt": 0.9, "force": False, "threshold": None,
                      "d_flex": {i: 0.1 * i for i in range(1, 11)}}
cd.structural_align = {"reference": "1AXC", "ref_label": "1AXC", "tm_ref": 0.90, "tm_query": 0.88,
                       "rmsd": 1.1, "n_aligned": 10, "norm": "ref", "matrix": [1, 0, 0, 0] * 3}

# V1 — folded + deviation + stability (real reducers)
v1 = cd.add_variant("V1"); cd.edit_variant("V1", 2, "W")
v1.results.fold = fold_summary({"engine": "boltz", "target": "monomer", "model_id": "3",
                                "mean_plddt": 89.0, "plddt": {i: 89.0 - i for i in range(1, 11)}}, author)
v1.results.fold["deviation"] = {"ddm": {str(i): 0.2 * i for i in range(1, 11)},
                                "lddt": {str(i): 0.99 - 0.02 * i for i in range(1, 11)},
                                "floor_ddm": {str(i): 0.3 for i in range(1, 11)},
                                "floor_lddt": {str(i): 0.9 for i in range(1, 11)}, "multichain": False}
v1.results.stability = stability_summary(
    [{"resnum": 3, "from_aa": "D", "to_aa": "W", "ddg": 1.2, "combined_score": 0.5, "recommendation": "ok"}],
    [type("M", (), {"resnum": 3, "to_aa": "W"})()])     # a (resnum,to_aa) match for V1

# V2 — solubility only
v2 = cd.add_variant("V2"); cd.edit_variant("V2", 1, "S")
v2.results.solubility = {"variant": 1.1, "wt": 1.0, "delta": 0.1}

# V3 — nothing
cd.add_variant("V3")

# an MPNN-imported variant via the REAL import path (provenance carries mpnn_run/design_k/fasta_path)
from variant_model import import_mpnn_designs
mpnn_seq = SEQ[:1] + "M" + SEQ[2:]                            # one substitution at resnum 2
mpnn_result = {"chain": "A", "fasta_path": "/tmp/run.fa",
               "sequences": [{"sequence": SEQ, "score": -0.5, "recovery": 1.0},
                             {"sequence": mpnn_seq, "score": -1.23, "recovery": 0.91}]}
_tmp = iter(range(1000))
for mv in import_mpnn_designs(cd, mpnn_result, run_id=0, next_id_fn=lambda: f"M{next(_tmp)}"):
    cd.variants.append(mv)
mpnn_vid = cd.variants[-1].id                                 # the design_k=1 import (mpnn_seq)

# ── persist + a named session folder ──
config.SESSION_DIR = Path(tempfile.mkdtemp(prefix="verify_export_sessions_"))
session = SessionState()
session.add_design_session(design.model_id, design.to_dict())
session.add_proteinmpnn_result(design.model_id, mpnn_result)
session_io.save_named_session(None, session, "mix")          # creates mix/ + folds/ + exports/

# ── EXPORT (with proteinmpnn_results so MPNN rows can carry the design score) ──
exports = session_io.session_paths("mix")["exports"]
rep = session_export.export_session(session.design_sessions, exports, session.proteinmpnn_results)
check("export wrote something (any)", rep["any"])
check("results.xlsx written", (exports / "results.xlsx").is_file())
import openpyxl
wb = openpyxl.load_workbook(str(exports / "results.xlsx"))
check("workbook has Summary + design-content + all populated result sheets",
      wb.sheetnames == ["Summary", "Sequences", "Substitutions", "Fold pLDDT", "Deviation",
                        "Stability ddG", "Solubility", "Template assist", "Template assist dflex",
                        "Structural align"])
import csv as _csv
def read_csv(name):
    with open(exports / "csv" / f"{name}.csv", newline="", encoding="utf-8") as fh:
        return list(_csv.reader(fh))
slug = session_export._SLUG
# sheet headers == csv headers (column parity) for every EMITTED sheet
title_key = {t: k for k, t, _s in session_export._TYPES}
parity = all([c.value for c in wb[t][1]] == read_csv(slug[title_key[t]])[0] for t in wb.sheetnames)
check("workbook sheet columns == CSV columns for every type", parity)

# ── phase 1.5: sequences + substitutions + FASTA, MPNN score joined, hand-made blank ──
seqrows = read_csv("sequences"); qh = seqrows[0]
Q = {r[qh.index("row")]: dict(zip(qh, r)) for r in seqrows[1:]}
check("Sequences sheet has T + variants with sequences", "T" in Q and "V1" in Q and mpnn_vid in Q)
check("MPNN variant sequence row carries the joined design score+recovery",
      Q[mpnn_vid]["source"] == "proteinmpnn" and Q[mpnn_vid]["mpnn_score"] == "-1.23"
      and Q[mpnn_vid]["mpnn_recovery"] == "0.91")
check("hand-made variant has blank MPNN score", Q["V1"]["mpnn_score"] == "")
subrows = read_csv("substitutions"); xh = subrows[0]
mpnn_subs = [r for r in subrows[1:] if r[xh.index("variant")] == mpnn_vid]
v1_subs = [r for r in subrows[1:] if r[xh.index("variant")] == "V1"]
check("MPNN substitution row carries source=proteinmpnn + score",
      bool(mpnn_subs) and mpnn_subs[0][xh.index("source")] == "proteinmpnn"
      and mpnn_subs[0][xh.index("score")] == "-1.23")
check("hand-made substitution row has source=manual + blank score",
      bool(v1_subs) and v1_subs[0][xh.index("source")] == "manual" and v1_subs[0][xh.index("score")] == "")
fasta = (exports / "sequences.fasta").read_text()
check("sequences.fasta written with T + MPNN-score header",
      f">{design.model_id}_A_T source=template" in fasta and "mpnn_score=-1.23" in fasta and SEQ in fasta)

# partial-data Summary: blanks, never 0
srows = read_csv("summary"); shdr = srows[0]
S = {r[shdr.index("row")]: dict(zip(shdr, r)) for r in srows[1:]}
check("Summary rows = T, V1, V2 (V3 with no results is absent)", set(S) == {"T", "V1", "V2"})
check("V1 joins fold+stab+dev", S["V1"]["mean_plddt"] == "89.0" and S["V1"]["sum_ddg"] == "1.2"
      and S["V1"]["max_dRMSD"] == "2.0")
check("V1 missing solubility is BLANK (not 0)", S["V1"]["solubility_delta"] == "")
check("V2 (solubility-only) has blank fold/stab/dev cells",
      S["V2"]["solubility_delta"] == "0.1" and S["V2"]["mean_plddt"] == "" and S["V2"]["sum_ddg"] == "")
check("construct adoption + tm_align present on rows", S["T"]["adoption"] == "0.93" and S["T"]["tm_align"] == "0.9")
# deviation keeps BOTH metrics + floors
drows = read_csv("deviation"); dhdr = drows[0]
check("deviation CSV has dRMSD + dRMSD_floor + lDDT + lDDT_floor columns",
      all(c in dhdr for c in ("dRMSD", "dRMSD_floor", "lDDT", "lDDT_floor")) and len(drows) == 11)

# ── phase 2: relevance-gated profile figures (plots for profiles, none for scalars) ──
figdir = exports / "figures"
figs = {p.name for p in figdir.glob("*.png")} if figdir.exists() else set()
check("pLDDT profile plot generated", any(n.startswith("plddt_") for n in figs))
check("deviation profile plot generated (dRMSD + lDDT)", any(n.startswith("deviation_") for n in figs))
check("template-assist Δflex profile plot generated", any(n.startswith("dflex_") for n in figs))
check("NO figure for scalar types (solubility/structural-align/stability)",
      not any(n.startswith(("solubility", "structural", "stability")) for n in figs))
check("figures are real PNGs (not empty placeholders)",
      all((figdir / n).stat().st_size > 1000 for n in figs) and bool(figs))

# ── SAVE AS → self-contained fork ──
fa = session_io.save_as_session(None, session, "mix_fork", src_name="mix")
check("Save As succeeded", fa["error"] is None)
fork = config.SESSION_DIR / "mix_fork"
check("fork inherited exports/ (results.xlsx travels)", (fork / "exports" / "results.xlsx").is_file())
check("fork has its own durable folds/ copy", any((fork / "folds").glob("*.cif")))
# open the fork independently — its json fold paths point INTO the fork
forked = SessionState.load(str(fork / "session.json"))
tf = forked.get_design_session(design.model_id)["chains"][next(iter(
    forked.get_design_session(design.model_id)["chains"]))]["template_fold"]
cifp = tf.get("cif_path")
check("fork session.json is self-contained (fold path inside the fork, or no fold file)",
      cifp is None or Path(cifp).parent == (fork / "folds"))
check("original 'mix' left frozen as a snapshot", (config.SESSION_DIR / "mix" / "session.json").is_file())

print(f"\n══ RESULT: {len(PASS)} passed, {len(FAIL)} failed ══")
if FAIL: print("FAILED:", FAIL); sys.exit(1)
print("DONE — exports (xlsx + CSVs) + Save As verified on a mixed real session.")
