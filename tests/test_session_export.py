"""
tests/test_session_export.py
----------------------------
The session EXPORTS core (session_export) — one row-building pass feeding BOTH the multi-tab
workbook and per-type CSVs, the partial-data Summary roll-up (blank cells, never 0/crash), and
FAIL-LOUD skip-empty (no header-only files; nothing-has-data → no files).
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

import session_export
from session_export import build_tables, export_session, _COLUMNS, _TITLE


def _mixed_sessions():
    """A MIXED session: a construct with template-assist + structural-align; V1 folded + deviation
    + stability; V2 solubility-only (partial); V3 no results at all."""
    return {"denovo-1": {"model_id": "denovo-1", "source": "sequence", "chains": {"c1": {
        "rep_chain": "A", "members": [["3", "A"]],
        "template_fold": {"engine": "boltz", "target": "monomer", "mean_plddt": 90.0,
                          "plddt": {1: 90.0, 2: 88.0}},
        "guided_fold": {},
        "template_assist": {"template_label": "8UB2", "unguided_mean_plddt": 92.0,
                            "guided_mean_plddt": 95.0, "d_plddt": 3.0, "n_stabilized": 2,
                            "n_residues": 2, "mean_d_flex": 0.1, "max_adoption": 0.93,
                            "tm_adopt": 0.9, "force": False, "threshold": None,
                            "d_flex": {1: 0.4, 2: 0.1}},
        "structural_align": {"reference": "1AXC", "ref_label": "1AXC", "tm_ref": 0.90,
                             "tm_query": 0.88, "rmsd": 1.1, "n_aligned": 2, "norm": "ref"},
        "variants": [
            {"id": "V1", "mutations": [{"resnum": 2, "from_aa": "C", "to_aa": "W", "source": "manual"}],
             "results": {"fold": {"engine": "boltz", "target": "monomer", "mean_plddt": 89.0,
                                  "plddt": {1: 89.0, 2: 87.0},
                                  "deviation": {"ddm": {"1": 0.5, "2": 2.0}, "lddt": {"1": 0.95, "2": 0.7},
                                                "floor_ddm": {"1": 0.3, "2": 0.3},
                                                "floor_lddt": {"1": 0.9, "2": 0.9}, "multichain": False}},
                         "stability": {"rows": [{"resnum": 2, "from_aa": "C", "to_aa": "W", "ddg": 1.2,
                                                 "ddg_source": "rosetta", "combined_score": 0.5,
                                                 "recommendation": "ok"}], "sum_ddg": 1.2, "tier": "deep"},
                         "solubility": None}},
            {"id": "V2", "mutations": [{"resnum": 1, "from_aa": "A", "to_aa": "S", "source": "manual"}],
             "results": {"fold": None, "stability": None,
                         "solubility": {"variant": 1.1, "wt": 1.0, "delta": 0.1}}},
            {"id": "V3", "mutations": [], "results": {"fold": None, "stability": None, "solubility": None}},
        ]}}}}


def _read_csv(p):
    with open(p, newline="", encoding="utf-8") as fh:
        return list(csv.reader(fh))


def test_export_writes_workbook_and_csvs_with_matching_columns(tmp_path):
    rep = export_session(_mixed_sessions(), tmp_path)
    assert rep["any"] is True
    xlsx = tmp_path / "results.xlsx"
    assert xlsx.is_file()

    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx))
    # Substitutions appears (variants have mutations); Sequences does NOT (this fixture has no cells)
    expected = ["Summary", "Substitutions", "Fold pLDDT", "Deviation", "Stability ddG",
                "Solubility", "Template assist", "Template assist dflex", "Structural align"]
    assert wb.sheetnames == expected
    # for every EMITTED sheet: workbook header == CSV header == _COLUMNS[key]
    title_to_key = {t: k for k, t, _s in session_export._TYPES}
    for title in wb.sheetnames:
        key = title_to_key[title]
        assert [c.value for c in wb[title][1]] == _COLUMNS[key]
        assert _read_csv(tmp_path / "csv" / f"{session_export._SLUG[key]}.csv")[0] == _COLUMNS[key]


def test_summary_partial_data_blank_never_zero(tmp_path):
    export_session(_mixed_sessions(), tmp_path)
    rows = _read_csv(tmp_path / "csv" / "summary.csv")
    hdr, data = rows[0], rows[1:]
    by_row = {r[hdr.index("row")]: dict(zip(hdr, r)) for r in data}
    assert set(by_row) == {"T", "V1", "V2"}             # V3 (no results) absent

    v1 = by_row["V1"]
    assert v1["mean_plddt"] == "89.0" and v1["sum_ddg"] == "1.2" and v1["max_dRMSD"] == "2.0"
    assert v1["solubility_delta"] == ""                 # missing → BLANK, not 0
    assert v1["adoption"] == "0.93" and v1["tm_align"] == "0.9"

    v2 = by_row["V2"]                                    # solubility-only variant
    assert v2["solubility_delta"] == "0.1"
    assert v2["mean_plddt"] == "" and v2["sum_ddg"] == "" and v2["max_dRMSD"] == ""

    t = by_row["T"]                                      # construct row: assist/align present
    assert t["mean_plddt"] == "90.0" and t["adoption"] == "0.93" and t["tm_align"] == "0.9"
    assert t["sum_ddg"] == ""


def test_deviation_keeps_both_metrics_and_floors(tmp_path):
    export_session(_mixed_sessions(), tmp_path)
    rows = _read_csv(tmp_path / "csv" / "deviation.csv")
    hdr, data = rows[0], rows[1:]
    assert hdr == _COLUMNS["deviation"]
    d = {r[hdr.index("resnum")]: dict(zip(hdr, r)) for r in data}
    assert d["2"]["dRMSD"] == "2.0" and d["2"]["dRMSD_floor"] == "0.3"
    assert d["2"]["lDDT"] == "0.7" and d["2"]["lDDT_floor"] == "0.9"


def test_remote_msa_provenance_reaches_fold_plddt_export(tmp_path):
    # A ColabFold (remote-MSA) construct fold → the export's fold_plddt rows MUST carry the
    # remote_msa tag (the column reaches the table, not just the badge), so a saved/exported
    # session never blurs which folds left LOCAL-ONLY. Local engines tag False.
    ds = {"dn-1": {"model_id": "dn-1", "source": "sequence", "chains": {"c1": {
        "rep_chain": "A", "members": [["2", "A"]],
        "template_fold": {"engine": "colabfold", "target": "monomer", "mean_plddt": 88.0,
                          "remote_msa": True, "plddt": {1: 88.0, 2: 90.0}},
        "variants": []}}}}
    assert "remote_msa" in _COLUMNS["fold_plddt"]
    tables = build_tables(ds)
    fp = tables["fold_plddt"]
    assert fp and all(row["remote_msa"] is True for row in fp)     # remote fold tagged in every row
    # a LOCAL Boltz fold tags remote_msa False (never blank/missing)
    ds["dn-1"]["chains"]["c1"]["template_fold"] = {
        "engine": "boltz", "target": "monomer", "mean_plddt": 90.0, "plddt": {1: 90.0}}
    fp2 = build_tables(ds)["fold_plddt"]
    assert fp2 and all(row["remote_msa"] is False for row in fp2)


def test_failloud_skips_empty_types(tmp_path):
    # a session with ONLY solubility data → deviation/stability/fold/assist/align all skipped
    ds = {"m": {"source": "sequence", "chains": {"c": {"rep_chain": "A", "variants": [
        {"id": "V1", "mutations": [], "results": {"solubility": {"variant": 1.0, "wt": 0.9, "delta": 0.1}}}]}}}}
    rep = export_session(ds, tmp_path)
    assert rep["any"] is True
    assert "Solubility" in rep["written"] and "Deviation" in rep["skipped"]
    assert not (tmp_path / "csv" / "deviation.csv").exists()      # no empty file
    import openpyxl
    wb = openpyxl.load_workbook(str(tmp_path / "results.xlsx"))
    assert wb.sheetnames == ["Summary", "Solubility"]


def test_nothing_to_export_writes_no_files(tmp_path):
    ds = {"m": {"source": "sequence", "chains": {"c": {"rep_chain": "A", "variants": [
        {"id": "V1", "mutations": [], "results": {}}]}}}}
    rep = export_session(ds, tmp_path)
    assert rep["any"] is False and rep["files"] == []
    assert not (tmp_path / "results.xlsx").exists()
    assert not (tmp_path / "csv").exists()


def test_build_tables_no_recompute_is_pure():
    ds = _mixed_sessions()
    import copy
    snap = copy.deepcopy(ds)
    build_tables(ds)
    assert ds == snap                                    # input not mutated


def _design_with_sequences():
    """Template ACDE; V1 hand-made C2W; V2 MPNN C2Y (joinable score); V3 MPNN whose stored
    sequence MISMATCHES (a later run overwrote) → score must blank, never wrong."""
    tcells = [{"col": i, "resnum": i + 1, "aa": a} for i, a in enumerate("ACDE")]
    def vc(seq): return [{"col": i, "resnum": i + 1, "aa": a} for i, a in enumerate(seq)]
    ds = {"denovo-1": {"model_id": "denovo-1", "source": "sequence", "chains": {"c": {
        "rep_chain": "A", "template_cells": tcells, "variants": [
            {"id": "V1", "source": "manual", "provenance": {}, "cells": vc("AWDE"),
             "mutations": [{"resnum": 2, "from_aa": "C", "to_aa": "W", "source": "manual"}],
             "indels": [], "results": {}},
            {"id": "V2", "source": "proteinmpnn",
             "provenance": {"mpnn_run": 0, "design_k": 1, "fasta_path": "/x.fa"}, "cells": vc("AYDE"),
             "mutations": [{"resnum": 2, "from_aa": "C", "to_aa": "Y", "source": "proteinmpnn"}],
             "indels": [], "results": {}},
            {"id": "V3", "source": "proteinmpnn",
             "provenance": {"mpnn_run": 0, "design_k": 0, "fasta_path": "/x.fa"}, "cells": vc("AHDE"),
             "mutations": [{"resnum": 2, "from_aa": "C", "to_aa": "H", "source": "proteinmpnn"}],
             "indels": [], "results": {}},
        ]}}}}
    # stored result: design_k 1 == AYDE (matches V2); design_k 0 == AZDE (does NOT match V3's AHDE)
    ppr = {"denovo-1": {"data": {"fasta_path": "/x.fa", "sequences": [
        {"sequence": "AZDE", "score": -1.0, "recovery": 0.80},
        {"sequence": "AYDE", "score": -1.23, "recovery": 0.91}]}}}
    return ds, ppr


def test_sequences_and_substitutions_with_mpnn_join(tmp_path):
    ds, ppr = _design_with_sequences()
    rep = export_session(ds, tmp_path, ppr)
    assert rep["any"] is True

    seq = _read_csv(tmp_path / "csv" / "sequences.csv"); sh = seq[0]
    S = {r[sh.index("row")]: dict(zip(sh, r)) for r in seq[1:]}
    assert S["T"]["sequence"] == "ACDE" and S["T"]["source"] == "template"
    assert S["V2"]["source"] == "proteinmpnn" and S["V2"]["mpnn_score"] == "-1.23" \
        and S["V2"]["mpnn_recovery"] == "0.91"
    assert S["V1"]["mpnn_score"] == ""                 # hand-made → blank
    assert S["V3"]["mpnn_score"] == ""                 # MPNN but stored seq mismatched → BLANK (not wrong)

    sub = _read_csv(tmp_path / "csv" / "substitutions.csv"); bh = sub[0]
    B = {r[bh.index("variant")]: dict(zip(bh, r)) for r in sub[1:]}
    assert B["V1"]["kind"] == "substitution" and B["V1"]["from_aa"] == "C" and B["V1"]["to_aa"] == "W"
    assert B["V1"]["source"] == "manual" and B["V1"]["score"] == ""        # hand-made → blank score
    assert B["V2"]["source"] == "proteinmpnn" and B["V2"]["score"] == "-1.23" and B["V2"]["recovery"] == "0.91"

    # FASTA travels, with the MPNN score in the header
    fasta = (tmp_path / "sequences.fasta").read_text()
    assert ">denovo-1_A_T source=template" in fasta and "ACDE" in fasta
    assert ">denovo-1_A_V2 source=proteinmpnn mpnn_score=-1.23" in fasta


def test_indels_appear_in_substitutions(tmp_path):
    ds = {"m": {"source": "sequence", "chains": {"c": {"rep_chain": "A",
        "template_cells": [{"col": i, "resnum": i + 1, "aa": a} for i, a in enumerate("ACDE")],
        "variants": [{"id": "V1", "source": "manual", "provenance": {},
                      "cells": [{"col": 0, "resnum": 1, "aa": "A"}, {"col": 1, "resnum": None, "aa": None}],
                      "mutations": [],
                      "indels": [{"kind": "deletion", "col": 1, "resnum": 2, "from_aa": "C"},
                                 {"kind": "insertion", "col": 4, "resnum": 4, "residues": "GG"}],
                      "results": {}}]}}}}
    export_session(ds, tmp_path)
    rows = _read_csv(tmp_path / "csv" / "substitutions.csv"); h = rows[0]
    kinds = {r[h.index("kind")] for r in rows[1:]}
    assert kinds == {"deletion", "insertion"}
    ins = [r for r in rows[1:] if r[h.index("kind")] == "insertion"][0]
    assert ins[h.index("residues")] == "GG"


def test_bare_construct_exports_its_sequence(tmp_path):
    # a de-novo construct with a sequence but NO results / variants → still exportable (design content)
    ds = {"denovo-1": {"source": "sequence", "chains": {"c": {"rep_chain": "A",
        "template_cells": [{"col": i, "resnum": i + 1, "aa": a} for i, a in enumerate("ACDEFG")],
        "variants": []}}}}
    rep = export_session(ds, tmp_path)
    assert rep["any"] is True and "Sequences" in rep["written"]
    assert (tmp_path / "sequences.fasta").is_file()
    import openpyxl
    assert openpyxl.load_workbook(str(tmp_path / "results.xlsx")).sheetnames == ["Sequences"]
